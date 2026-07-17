import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE_DIR = Path(__file__).resolve().parents[2]

DATABASE = BASE_DIR / "database" / "financial_data.db"

OUTPUT = BASE_DIR / "output"

OUTPUT.mkdir(exist_ok=True)

connection = sqlite3.connect(DATABASE)

financial = pd.read_sql(
    "SELECT * FROM financial_ratios",
    connection
)

peer = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    connection
)

connection.close()

print(financial.shape)
print(peer.shape)


groups = peer["peer_group"].unique()

print(groups)

output_file = OUTPUT / "peer_comparison.xlsx"

writer = pd.ExcelWriter(
    output_file,
    engine="openpyxl"
)

for group in groups:

    companies = peer[
        peer["peer_group"] == group
    ]["company_id"].unique()

    report = financial[
        financial["company_id"].isin(companies)
    ]

    report.to_excel(
        writer,
        sheet_name=group[:31],
        index=False
    )

writer.close()

print("Peer Comparison Report Created!")


wb = load_workbook(output_file)

green = PatternFill(
    fill_type="solid",
    start_color="90EE90"
)

yellow = PatternFill(
    fill_type="solid",
    start_color="FFFF99"
)

red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)


for sheet in wb.sheetnames:

    ws = wb[sheet]

    headers = [
        cell.value
        for cell in ws[1]
    ]

    if "roe_calculated" not in headers:
        continue

    col = headers.index(
        "roe_calculated"
    ) + 1

    values = []

    for row in range(2, ws.max_row + 1):

        val = ws.cell(
            row=row,
            column=col
        ).value

        if isinstance(val, (int, float)):
            values.append(val)

    if len(values) == 0:
        continue

    p25 = pd.Series(values).quantile(0.25)

    p75 = pd.Series(values).quantile(0.75)

    for row in range(2, ws.max_row + 1):

        cell = ws.cell(
            row=row,
            column=col
        )

        value = cell.value

        if not isinstance(value, (int, float)):
            continue

        if value >= p75:
            cell.fill = green

        elif value <= p25:
            cell.fill = red

        else:
            cell.fill = yellow

wb.save(output_file)

print("Formatting Applied!")